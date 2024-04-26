from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def scrape_data(url, driver):
    
    driver.get(url)

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, "body")))

    scraped_data = {}
    scraped_data['text'] = driver.find_elements(By.XPATH, "//p | //span | //div")  
    scraped_data['images'] = driver.find_elements(By.TAG_NAME, "img")  
    scraped_data['links'] = driver.find_elements(By.TAG_NAME, "a") 

    text_content = ""
    for element in scraped_data['text']:
        text_content += element.text + "\n"
    scraped_data['text'] = text_content.strip()

    image_urls = []
    for image in scraped_data['images']:
        image_url = image.get_attribute("src")
        if image_url:
            image_urls.append(image_url)
    scraped_data['images'] = image_urls

    link_urls = []
    for link in scraped_data['links']:
        link_url = link.get_attribute("href")
        if link_url:
            link_urls.append(link_url)
    scraped_data['links'] = link_urls

    return scraped_data

def main():
    options = webdriver.ChromeOptions()
    options.add_argument("--headless") 
    driver = webdriver.Chrome()

    website_links = []
    try:
        workbook = load_workbook("C:\\Users\\ACER\\Desktop\\Web Scraper selenium\\LinksSeleniumWEBSC.xlsx")
        worksheet = workbook["Sheet1"]  
        for row in worksheet.iter_rows(min_row=2):  
            cell_value = row[0].value 
            if cell_value:
                website_links.append(cell_value)
    except FileNotFoundError:
        print("Error: Excel file not found!")
        return
    except Exception as e:
        print("Error reading Excel file:", e)
        return

    scraped_data = []
    for link in website_links:
        try:
            data = scrape_data(link, driver)
            scraped_data.append(data)
        except Exception as e:
            print(f"Error scraping {link}:", e)

    driver.quit()

    try:
        workbook = load_workbook("C:\\Users\\ACER\\Desktop\\Web Scraper selenium\\link2.xlsx")
        worksheet = workbook.create_sheet("Scraped_Data")
        
        headers = ['Text', 'Images', 'Links']
        for col, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=col, value=header)

        for row, data in enumerate(scraped_data, start=2):
            worksheet.cell(row=row, column=1, value=data['text'])
            worksheet.cell(row=row, column=2, value='\n'.join(data['images']))
            worksheet.cell(row=row, column=3, value='\n'.join(data['links']))

        workbook.save("C:\\Users\\ACER\\Desktop\\Web Scraper selenium\\link2.xlsx")
        print("Scraped data has been successfully written to the Excel file.")
    except FileNotFoundError:
        print("Error: Excel file not found!")
    except Exception as e:
        print("Error writing to Excel file:", e)

if __name__ == "__main__":
    main()
